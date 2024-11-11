import pysftp
import re
import datetime
import stat
import os
import threading
import queue
import gzip
import multiprocessing 
import xml.etree.cElementTree as ET
import sqlite3
import openpyxl
import shutil
import enmscripting
from concurrent.futures import ThreadPoolExecutor
import paramiko
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from pprint import pprint
import sys

def get_path_date(dt:datetime.datetime,template_datetime_str:str) -> str:

    minute_delta = 7

    datetime_str = ""

    delta_1 = datetime.timedelta(hours=1)

    if (45 + minute_delta) <= dt.minute and dt.minute < 60:
        dt_after  = dt

        datetime_str = template_datetime_str.format(y1=dt_after.strftime("%Y"), m1=dt_after.strftime("%m"), d1=dt_after.strftime("%d"), H1=dt_after.strftime("%H"), M1="30", H2=dt_after.strftime("%H"), M2="45")

    if 0 <= dt.minute and dt.minute < (0 + minute_delta):
        dt_after  = dt - delta_1

        datetime_str = template_datetime_str.format(y1=dt_after.strftime("%Y"), m1=dt_after.strftime("%m"), d1=dt_after.strftime("%d"), H1=dt_after.strftime("%H"), M1="30", H2=dt_after.strftime("%H"), M2="45")

    elif (0 + minute_delta) <= dt.minute and dt.minute < (15 + minute_delta):
        dt_after  = dt - delta_1

        datetime_str = template_datetime_str.format(y1=dt_after.strftime("%Y"), m1=dt_after.strftime("%m"), d1=dt_after.strftime("%d"), H1=dt_after.strftime("%H"), M1="45", H2=(dt_after + delta_1).strftime("%H"), M2="00")


    elif (15 + minute_delta) <= dt.minute and dt.minute < (30 + minute_delta):
        dt_after  = dt

        datetime_str = template_datetime_str.format(y1=dt_after.strftime("%Y"), m1=dt_after.strftime("%m"), d1=dt_after.strftime("%d"), H1=dt_after.strftime("%H"), M1="00", H2=dt_after.strftime("%H"), M2="15")


    elif (30 + minute_delta) <= dt.minute and dt.minute < (45 + minute_delta):
        dt_after  = dt 

        datetime_str = template_datetime_str.format(y1=dt_after.strftime("%Y"), m1=dt_after.strftime("%m"), d1=dt_after.strftime("%d"), H1=dt_after.strftime("%H"), M1="15", H2=dt_after.strftime("%H"), M2="30")

    return datetime_str

def get_report_path(dt:datetime.datetime, path:str) -> str:
    report_path = os.getcwd() + "\\report\\" + dt.strftime("%Y") + "\\" + dt.strftime("%m") + "\\" + dt.strftime("%d") + "\\" + dt.strftime("%H%M") + "\\" + path
    if not os.path.isdir(report_path):
        os.makedirs(report_path)
    return report_path

def get_file_list(host_ip,username,password,url,datetime_str) -> list:
    all_dir_list = list()

    cnopts = pysftp.CnOpts()
    cnopts.hostkeys = None
    with pysftp.Connection(host=host_ip, username=username, password=password,cnopts=cnopts) as sftp_connections:

        template_base = "{datetime_str}{dir_name}_statsfile.xml.gz"
        path_str = "{du_type}|{file_name}|{path}/{dir_name}/{file_name}"

        check_dirs = ["/ericsson/pmic1/XML","/ericsson/pmic2/XML"]

        for check_dir in check_dirs:
            sftp_connections.chdir(check_dir)
            dir_list = sftp_connections.listdir_attr()
            path = sftp_connections.getcwd()

            for item in dir_list:
                find_1 = re.search(r"SubNetwork",item.filename)
                find_2 = re.search(r"ManagedElement",item.filename)
                find_3 = re.search(r"PICO",item.filename)
                find_4 = stat.S_ISDIR(item.st_mode)
                find_5 = re.search(r"RNCE-MOS",item.filename)

                if find_1 and find_2 and not find_3 and find_4:
                    file_name = template_base.format(datetime_str=datetime_str, dir_name=item.filename)
                    all_dir_list.append(path_str.format(du_type="BB",file_name=file_name,path=path,dir_name=item.filename))
                
                elif find_1 and find_5 and not find_3 and find_4:
                    file_name = template_base.format(datetime_str=datetime_str, dir_name=item.filename)
                    all_dir_list.append(path_str.format(du_type="DUW",file_name=file_name,path=path,dir_name=item.filename))
                


    return all_dir_list

def enm_file_get(host_ip,username,password,url,report_path,file_list_str,queue_of_threads,enm):

    file_list = file_list_str.split("||")
    try:
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None
        with pysftp.Connection(host=host_ip, username=username, password=password,cnopts=cnopts) as sftp_connections:

            for items in file_list:

                if len(items.split("|")) == 3:
                    du_type,file_name,base_path = items.split("|")
                    
                    if sftp_connections.exists(base_path):
                        sftp_connections.get(base_path,report_path + "\\" + file_name)
                        queue_of_threads.put(enm + "|" + du_type + "|" + report_path + "\\" + file_name)
                    else:
                        pass
                else:
                    print(items.split("|"))

    except Exception as e:
        print(e)
        
def enm_connection_information() -> dict:

    enm_sftp_connection_dict = {"enm7":{"host_ip":"10.12.110.27","username":"1234","password":"1234","url":"https://enm7.enm.t*2.ru/"},    #Проверить правильность url
                                "enm8":{"host_ip":"10.12.112.25","username":"1234","password":"1234","url":"https://enm8.enm.t*2.ru/"},    #Проверить правильность url
                               "enm14":{"host_ip":"10.12.240.27","username":"1234","password":"1234","url":"https://enm14.enm.t*2.ru/"}}   #Проверить правильность url
    return enm_sftp_connection_dict

def enm_file_get_in_threads(threads_count,dt:datetime.datetime):

    enm_sftp_connection_dict = enm_connection_information()

    template_datetime_str = "A{y1}{m1}{d1}.{H1}{M1}+0300-{H2}{M2}+0300_"

    datetime_str = get_path_date(dt,template_datetime_str)
    print(datetime_str)
    threds_list = list()

    queue_of_threads = queue.Queue()

    path_list = list()

    for enm,connection_data in enm_sftp_connection_dict.items():
        report_path = get_report_path(dt,enm)
        
        file_list = get_file_list(**connection_data,datetime_str=datetime_str)

        whole_part = len(file_list) // threads_count


        remainder_of_part = len(file_list) % threads_count
        
        for i in range(threads_count):
            connection_dict = dict()
            connection_dict["report_path"] = report_path
            connection_dict["enm"] = enm
            connection_dict["queue_of_threads"] = queue_of_threads

            if i == threads_count-1 and remainder_of_part != 0:
                connection_dict["file_list_str"] = "||".join(file_list[threads_count-1*whole_part:])
            else:
                connection_dict["file_list_str"] = "||".join(file_list[i*whole_part:(i+1)*whole_part])

            connection_dict.update(connection_data)
            thread_dict = dict()
            thread_dict["enm"] = enm
            thread_dict["thread_nubmer"] = str(i+1)
            thread_dict["thread"] = threading.Thread(target=enm_file_get,kwargs=connection_dict)

            threds_list.append(thread_dict)

    for item in threds_list:
        item["thread"].start()
        #print("{enm} thread {thread_nubmer} has started".format(enm=item["enm"],thread_nubmer=item["thread_nubmer"]))

    for item in threds_list:
        item["thread"].join()
        #print("{enm} thread {thread_nubmer} has finished".format(enm=item["enm"],thread_nubmer=item["thread_nubmer"]))

    while not queue_of_threads.empty():
        path_list.append(queue_of_threads.get())

    return path_list

def open_gz(file_path:str) -> bytes:
    file_content = bytes()
    with gzip.open(file_path, 'rb') as f:
        file_content = f.read()
    return file_content

def get_temp_bb_find(file_content:bytes) -> dict:
    pmSfpTemperature = list()
    pmUnitTemperatureLevel = list()

    temp_dict = dict()
    find_dict = dict()
    equipment_temp_list = list()

    s = file_content.decode(encoding="utf-8")

    s_f = "<managedElement localDn=\""
    find1 = s.find("<managedElement localDn=\"")
    find2 = s.find("\"",find1+len(s_f))
    if find1 and find2:
        find_dict["ne"] = s[find1 + len(s_f):find2]

    if "ne" in find_dict and find_dict["ne"] != None:
        temp_dict["ne"] = find_dict["ne"]


    s_f = "<measInfo measInfoId=\"PM=1,PmGroup=FieldReplaceableUnit\">"
    find1 = s.find("<measInfo measInfoId=\"PM=1,PmGroup=FieldReplaceableUnit\">")
    find2 = s.find("</measInfo>",find1+len(s_f))
    if find1 and find2:
        find_dict["FieldReplaceableUnit"] = s[find1 + len(s_f):find2]

    if "FieldReplaceableUnit" in find_dict and find_dict["FieldReplaceableUnit"] != None:
        pointer1 = 0
        find_result = 1
        result_list = list()
        temp = find_dict["FieldReplaceableUnit"]
        while find_result == 1:
            find_result = 0
            temp = temp[pointer1:]
            find_p1 = temp.find("<measValue")
            if find_p1 != -1:
                find_p2 = temp.find(">",find_p1 + len("<measValue"))
                if find_p2 != -1:
                    find_p3 = temp.find("</measValue>",find_p2 + len("<"))
                    if find_p3 != -1:
                        pointer1 = find_p3 + len("</measValue>")
                        if pointer1 >= len(temp) - 1:
                            find_result = 0
                        else:
                            find_result = 1

                        find_bb = temp.find("FieldReplaceableUnit=BB-1",find_p1 + len("<measValue"),find_p2)
                        find_p4 = temp.find("<r p=\"2\">",find_p2 + len(">"),find_p3)
                        if find_p4 != -1 and find_bb != -1:

                            find_p5 = temp.find("</r>",find_p4 + len("<r p=\"2\">"),find_p3)
                            if find_p5 != -1:
                                result_list.append(temp[find_p4 + len("<r p=\"2\">"):find_p5])


        temperature_level = list()
        if result_list != []:
            temperature_level = result_list[0].replace(" ","").split(",")

        if len(temperature_level) != 0:
            for item in temperature_level:
                if item.isdigit():
                    equipment_temp_list.append(int(item))
            if equipment_temp_list != []:
                pmUnitTemperatureLevel.append(round(sum(equipment_temp_list)/len(equipment_temp_list)))


    s_f = "<measInfo measInfoId=\"PM=1,PmGroup=SfpModule\">"
    find1 = s.find("<measInfo measInfoId=\"PM=1,PmGroup=SfpModule\">")
    find2 = s.find("</measInfo>",find1+len(s_f))
    if find1 and find2:
        find_dict["SfpModule"] = s[find1 + len(s_f):find2]


    if "SfpModule" in find_dict and find_dict["SfpModule"] != None:
        pointer1 = 0
        find_result = 1
        result_list = list()
        temp = find_dict["SfpModule"]
        while find_result == 1:
            find_result = 0
            temp = temp[pointer1:]
            find_p1 = temp.find("<measValue")
            if find_p1 != -1:
                find_p2 = temp.find(">",find_p1 + len("<measValue"))
                if find_p2 != -1:
                    
                    find_p3 = temp.find("</measValue>",find_p2 + len("<"))
                    if find_p3 != -1:

                        pointer1 = find_p3 + len("</measValue>")
                        if pointer1 >= len(temp) - 1:
                            find_result = 0
                        else:
                            find_result = 1

                        find_bb = temp.find("FieldReplaceableUnit=BB-1",find_p1 + len("<measValue"),find_p2)
                        find_p4 = temp.find("<r p=\"1\">",find_p2 + len(">"),find_p3)
                        if find_p4 != -1 and find_bb != -1:

                            find_p5 = temp.find("</r>",find_p4 + len("<r p=\"1\">"),find_p3)
                            if find_p5 != -1:
                                result_list.append(temp[find_p4 + len("<r p=\"1\">"):find_p5])

        if result_list != []:
            for item in result_list:
                temperature = item.replace(" ","")
                if temperature != "" and temperature != "0" and temperature.isdigit():
                    pmSfpTemperature.append(round((int(temperature) - 1000))/10)


    if pmSfpTemperature != []:
        temp_dict["temperature"] = min(pmSfpTemperature)
    else:
        temp_dict["temperature"] = None
                        
    if pmUnitTemperatureLevel != []:
        temp_dict["temperature_level"] = round(sum(pmUnitTemperatureLevel)/len(pmUnitTemperatureLevel))
    else:
        temp_dict["temperature_level"] = None

    return temp_dict

def get_temp_bb_etree(file_content:bytes) -> dict:

    pmSfpTemperature = list()
    pmUnitTemperatureLevel = list()

    temp_dict = dict()

    equipment_temp_list = list()

    root = ET.fromstring(file_content)

    file_dict = dict()
    file_dict["ne"] = root[0][0].attrib["localDn"]
    file_dict["SfpModule"] = root.find("./{*}measData/*[@measInfoId='PM=1,PmGroup=SfpModule']")
    file_dict["FieldReplaceableUnit"] = root.find("./{*}measData/*[@measInfoId='PM=1,PmGroup=FieldReplaceableUnit']")

    if "ne" in file_dict and file_dict["ne"] != None:
        temp_dict["ne"] = file_dict["ne"]

    if "SfpModule" in file_dict and file_dict["SfpModule"] != None:
        names = file_dict["SfpModule"]

        for name in names:
            if name.find("./[@measObjLdn]"):
                if re.search("BB",name.attrib["measObjLdn"]):
                    temperature = name.find("./*[@p='1']").text.replace(" ","")
                    if temperature != "" and temperature != "0" and temperature.isdigit():
                        pmSfpTemperature.append(round((int(temperature) - 1000))/10)

    if "FieldReplaceableUnit" in file_dict and file_dict["FieldReplaceableUnit"] != None:
        names = file_dict["FieldReplaceableUnit"]

        for name in names:
            equipment_temp_list = list()

            if name.find("./[@measObjLdn]"):
                if re.search("BB",name.attrib["measObjLdn"]):
                    temperature_level = name.find("./*[@p='2']").text.replace(" ","").split(",")

                    if len(temperature_level) != 0:
                        for item in temperature_level:
                            if item.isdigit():
                                equipment_temp_list.append(int(item))
                        if equipment_temp_list != []:
                            pmUnitTemperatureLevel.append(round(sum(equipment_temp_list)/len(equipment_temp_list)))

    if pmSfpTemperature != []:
        temp_dict["temperature"] = min(pmSfpTemperature)
    else:
        temp_dict["temperature"] = None
                        
    if pmUnitTemperatureLevel != []:
        temp_dict["temperature_level"] = round(sum(pmUnitTemperatureLevel)/len(pmUnitTemperatureLevel))
    else:
        temp_dict["temperature_level"] = None

    return temp_dict

def get_temp_duw_etree(file_content:bytes) -> dict:

    temp_dict = dict()

    equipment_temp_list = list()

    root = ET.fromstring(file_content)

    file_dict = dict()

    if root.find("./md/neid/nedn") is not None:
        find_1 = re.search(r"=[MSCOR]{2}\d{4}",root.find("./md/neid/nedn").text)
        if find_1:
            file_dict["ne"] = (find_1.group(0)[1:].upper())

    if root.find("./md/mi") is not None:
        for item in root.find("./md/mi"):
            if item.text == "pmCabinetTemperature":
                if len(root.find("./md/mi/mv"))==6:
                    file_dict["pmCabinetTemperature"] = root.find("./md/mi/mv")[5].text

    if "ne" in file_dict and len(file_dict["ne"]) == 6:
        temp_dict["ne"] = file_dict["ne"]


    if "pmCabinetTemperature" in file_dict and file_dict["pmCabinetTemperature"] != None:

        temperature_level = file_dict["pmCabinetTemperature"].replace(" ","").split(",")
        
        if len(temperature_level) != 0:
            for item in temperature_level:
                if item.isdigit():
                    equipment_temp_list.append(int(item) - 1000)

            if equipment_temp_list != []:
                temp_dict["temperature"] = round(sum(equipment_temp_list)/len(equipment_temp_list))
                temp_dict["temperature_level"] = None

    return temp_dict

def file_execution(file_str:str) -> dict:

    result_dict = dict()
    temp_path = file_str.split("|")

    if len(temp_path) == 3:
        enm,du_type,path = temp_path

        result_xml = bytes()
        dict_xml = dict()

        try:
            result_xml = open_gz(path)
        except Exception as e:
            #result_error_list.append("gz_error " + path + " " + str(e))
            print("gz_error " + path + " " + str(e))

        if result_xml != "":
            if du_type == "BB":
                result_dict["enm"] = enm
                result_dict["du_type"] = du_type
                try:
                    dict_xml = get_temp_bb_find(result_xml)
                except Exception as e:
                    #result_error_list.append("xml_error " + path + " " + str(e))
                    print("find_error " + path + " " + str(e))

            elif du_type == "DUW":
                result_dict["enm"] = enm
                result_dict["du_type"] = du_type
                try:
                    dict_xml = get_temp_duw_etree(result_xml)
                except Exception as e:
                    #result_error_list.append("xml_error " + path + " " + str(e))
                    print("xml_error " + path + " " + str(e))

            result_dict.update(dict_xml)

    return result_dict
    
def file_execution_process_wrap(file_list_str,process_queue):

    for i,item in enumerate(file_list_str):
        result_dict = dict()
        result_dict = file_execution(item)
        #print(str(i) + " " + str(result_dict))
        if result_dict != {}:
            if "error" in result_dict:
                print(result_dict)
            else:
                if "enm" in result_dict and "du_type" in result_dict and "ne" in result_dict and "temperature" in result_dict and "temperature_level" in result_dict:
                    process_queue.put(result_dict)

def file_execution_on_process(process_count:int,path_list:list) -> list:

    process_list = list()

    general_queue = multiprocessing.Manager().Queue()

    whole_part = len(path_list) // process_count
        
    for i in range(process_count):
        connection_dict = dict()
        connection_dict["process_queue"] = general_queue

        if i == process_count-1:
            connection_dict["file_list_str"] = path_list[i*whole_part:]

        else:
            connection_dict["file_list_str"] = path_list[i*whole_part:(i+1)*whole_part]


        process_dict = dict()

        process_dict["process_nubmer"] = str(i+1)

        process_dict["process"] = multiprocessing.Process(target=file_execution_process_wrap,kwargs=connection_dict)

        process_list.append(process_dict)

    for item in process_list:
        item["process"].start()
        print("process {process_nubmer} has started".format(process_nubmer=item["process_nubmer"]))

    for item in process_list:
        item["process"].join()
        print("process {process_nubmer} has finished".format(process_nubmer=item["process_nubmer"]))

    temp_list = list()

    print(general_queue.qsize())
    while not general_queue.empty():
        temp_list.append(general_queue.get())

    return temp_list

def addind_to_db(db_list: list, dt:datetime.datetime) -> list:

    result_list = list()

    template_datetime_str = "{y1}-{m1}-{d1} {H1}:{M1}"

    datetime_str = get_path_date(dt,template_datetime_str)

    sql_query1 = '''CREATE TABLE IF NOT EXISTS date_time_request (
                                                              id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                       date_time DATE UNIQUE);'''
    
    sql_query2 = '''SELECT id
                      FROM date_time_request
                     WHERE date_time = "{}";'''
    
    sql_query3 = '''INSERT INTO date_time_request (date_time)
                    VALUES ("{}");'''

    sql_query4 = '''CREATE TABLE IF NOT EXISTS sfp_temp (
                                                     id INTEGER PRIMARY KEY AUTOINCREMENT,
                                           date_time_id INTEGER NOT NULL,
                                                    enm TEXT NOT NULL,
                                                     bs TEXT NOT NULL,
                                                du_type TEXT NOT NULL,
                                                     ne TEXT NOT NULL,
                                            temperature INT,
                                      temperature_level INT);'''
    
    sql_query5 = '''INSERT INTO sfp_temp (
                            date_time_id,
                                     enm,
                                      bs,
                                 du_type,
                                      ne,
                             temperature,
                       temperature_level)

                    VALUES ({},:enm,:bs,:du_type,:ne,:temperature,:temperature_level);'''


    sql_query6 = '''CREATE TABLE IF NOT EXISTS enm_bs_ne (
                                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                                           enm TEXT NOT NULL,
                                            bs TEXT NOT NULL,
                                            ne TEXT NOT NULL);'''

    sql_query6_1 = '''INSERT INTO enm_bs_ne (enm, bs, ne)
                                   SELECT TEMP3.enm, TEMP3.bs, TEMP3.ne
                                     FROM
                                          (SELECT TEMP2.enm, TEMP2.bs, TEMP2.ne, enm_bs_ne.id
                                             FROM 
                                                  (SELECT DISTINCT enm, bs, ne
                                                     FROM sfp_temp
                                                    WHERE date_time_id = {}) AS TEMP2
                                        LEFT JOIN enm_bs_ne
                                               ON TEMP2.enm = enm_bs_ne.enm AND TEMP2.bs = enm_bs_ne.bs AND TEMP2.ne = enm_bs_ne.ne) AS TEMP3
                                    WHERE TEMP3.id IS NULL;'''

    sql_query7 = '''CREATE TABLE IF NOT EXISTS enm_bs (
                                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                                           enm TEXT NOT NULL,
                                            bs TEXT NOT NULL);'''

    sql_query7_1 = '''INSERT INTO enm_bs (enm, bs)
                                   SELECT TEMP3.enm, TEMP3.bs
                                     FROM
                                          (SELECT TEMP2.enm, TEMP2.bs, enm_bs.id
                                             FROM 
                                                  (SELECT DISTINCT enm, bs
                                                     FROM sfp_temp
                                                    WHERE date_time_id = {}) AS TEMP2
                                        LEFT JOIN enm_bs
                                               ON TEMP2.enm = enm_bs.enm AND TEMP2.bs = enm_bs.bs) AS TEMP3
                                    WHERE TEMP3.id IS NULL;'''

    sql_query8_1 = '''DROP TABLE IF EXISTS final_table;'''

    sql_query8= '''CREATE TABLE IF NOT EXISTS final_table
                                        AS
                                    SELECT enm_bs_ne.id, enm_bs.id as group_id, sfp_temp.bs, bs_reg.reg,sfp_temp.enm,sfp_temp.ne,sfp_temp.du_type,sfp_temp.temperature,sfp_temp.temperature_level
                                      FROM sfp_temp
                                 LEFT JOIN enm_bs_ne
                                        ON sfp_temp.enm = enm_bs_ne.enm AND sfp_temp.bs = enm_bs_ne.bs AND sfp_temp.ne = enm_bs_ne.ne 
                                 LEFT JOIN enm_bs
                                        ON sfp_temp.enm = enm_bs.enm AND sfp_temp.bs = enm_bs.bs
                                 LEFT JOIN bs_reg
                                        ON sfp_temp.bs = bs_reg.bs
                                     WHERE sfp_temp.date_time_id = {};'''

    sql_query9 = '''CREATE TEMPORARY TABLE TEMP_temperature (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT)'''

    sql_query10 = '''INSERT INTO TEMP_temperature (group_id,reg)
                                           SELECT group_id,reg
                                             FROM final_table
                                         GROUP BY group_id
                                         ORDER BY MAX(temperature) DESC;'''

    sql_query11 = '''CREATE TEMPORARY TABLE TEMP_temperature_level_bb (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query12_1 = '''INSERT INTO TEMP_temperature_level_bb (group_id,reg)
                                                      SELECT group_id,reg
                                                        FROM final_table
                                                       WHERE du_type = "BB"
                                                    GROUP BY group_id
                                                    ORDER BY MAX(temperature_level) DESC;'''

    sql_query12_2 = '''INSERT INTO TEMP_temperature_level_bb (group_id,reg)
                                                      SELECT group_id,reg
                                                        FROM
                                                     (SELECT group_id,reg, bs
                                                        FROM
                                                     (SELECT group_id,reg,bs
                                                        FROM final_table
                                                    GROUP BY group_id
                                                    ORDER BY bs)
                                                      EXCEPT
                                                      SELECT group_id,reg, bs
                                                        FROM 
                                                     (SELECT group_id,reg,bs
                                                        FROM final_table
                                                       WHERE du_type = "BB"
                                                    GROUP BY group_id
                                                    ORDER BY MAX(temperature_level) DESC))
                                                    ORDER BY bs;'''

    sql_query13 = '''CREATE TEMPORARY TABLE TEMP_temperature_bb (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query14_1 = '''INSERT INTO TEMP_temperature_bb (group_id,reg)
                                                 SELECT group_id,reg
                                                   FROM final_table
                                                  WHERE du_type = "BB"
                                               GROUP BY group_id
                                               ORDER BY MAX(temperature) DESC;'''

    sql_query14_2 = '''INSERT INTO TEMP_temperature_bb (group_id,reg)
                                                 SELECT group_id,reg
                                                   FROM
                                                (SELECT group_id,reg, bs
                                                   FROM
                                                (SELECT group_id,reg,bs
                                                   FROM final_table
                                               GROUP BY group_id
                                               ORDER BY bs)
                                                 EXCEPT
                                                 SELECT group_id,reg, bs
                                                   FROM 
                                                (SELECT group_id,reg,bs
                                                   FROM final_table
                                                  WHERE du_type = "BB"
                                               GROUP BY group_id
                                               ORDER BY MAX(temperature) DESC))
                                               ORDER BY bs;'''

    sql_query15 = '''CREATE TEMPORARY TABLE TEMP_temperature_duw (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query16_1 = '''INSERT INTO TEMP_temperature_duw (group_id,reg)
                                                 SELECT group_id,reg
                                                   FROM final_table
                                                  WHERE du_type = "DUW"
                                               GROUP BY group_id
                                               ORDER BY MAX(temperature) DESC;'''

    sql_query16_2 = '''INSERT INTO TEMP_temperature_duw (group_id,reg)
                                                 SELECT group_id,reg
                                                   FROM
                                                (SELECT group_id,reg, bs
                                                   FROM
                                                (SELECT group_id,reg,bs
                                                   FROM final_table
                                               GROUP BY group_id
                                               ORDER BY bs)
                                                 EXCEPT
                                                 SELECT group_id,reg, bs
                                                   FROM 
                                                (SELECT group_id,reg,bs
                                                   FROM final_table
                                                  WHERE du_type = "DUW"
                                               GROUP BY group_id
                                               ORDER BY MAX(temperature) DESC))
                                               ORDER BY bs;'''

    sql_query17_1 = '''CREATE TEMPORARY TABLE TEMP_temperature_tcu (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query17_2 = '''INSERT INTO TEMP_temperature_tcu (group_id,reg)
                                                 SELECT group_id,reg
                                                   FROM final_table
                                                  WHERE du_type = "TCU"
                                               GROUP BY group_id
                                               ORDER BY MAX(temperature) DESC;'''

    sql_query17_3 = '''INSERT INTO TEMP_temperature_tcu (group_id,reg)
                                                 SELECT group_id,reg
                                                   FROM
                                                (SELECT group_id,reg, bs
                                                   FROM
                                                (SELECT group_id,reg,bs
                                                   FROM final_table
                                               GROUP BY group_id
                                               ORDER BY bs)
                                                 EXCEPT
                                                 SELECT group_id,reg, bs
                                                   FROM 
                                                (SELECT group_id,reg,bs
                                                   FROM final_table
                                                  WHERE du_type = "TCU"
                                               GROUP BY group_id
                                               ORDER BY MAX(temperature) DESC))
                                               ORDER BY bs;'''

    sql_query18_1 = '''CREATE TEMPORARY TABLE TEMP_temperature_dus (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query18_2 = '''INSERT INTO TEMP_temperature_dus (group_id,reg)
                                                 SELECT group_id,reg
                                                   FROM final_table
                                                  WHERE du_type = "DUS"
                                               GROUP BY group_id
                                               ORDER BY MAX(temperature) DESC;'''

    sql_query18_3 = '''INSERT INTO TEMP_temperature_dus (group_id,reg)
                                                 SELECT group_id,reg
                                                   FROM
                                                (SELECT group_id,reg, bs
                                                   FROM
                                                (SELECT group_id,reg,bs
                                                   FROM final_table
                                               GROUP BY group_id
                                               ORDER BY bs)
                                                 EXCEPT
                                                 SELECT group_id,reg, bs
                                                   FROM 
                                                (SELECT group_id,reg,bs
                                                   FROM final_table
                                                  WHERE du_type = "DUS"
                                               GROUP BY group_id
                                               ORDER BY MAX(temperature) DESC))
                                               ORDER BY bs;'''

    sql_query24 = '''CREATE TEMPORARY TABLE TEMP_temperature_1 (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query24_1 = '''INSERT INTO TEMP_temperature_1 (group_id, reg)
                                               SELECT group_id,reg
                                                 FROM
                                              (SELECT
                                                      id,
                                                      group_id,
                                                      reg,
                                                      row_number() OVER (PARTITION BY reg ORDER BY id) g_f
                                                 FROM TEMP_temperature);'''

    sql_query25 = '''CREATE TEMPORARY TABLE TEMP_temperature_level_bb_1 (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query25_1 = '''INSERT INTO TEMP_temperature_level_bb_1 (group_id, reg)
                            SELECT group_id,reg
                              FROM
                           (SELECT
                                   id,
                                   group_id,
                                   reg,
                                   row_number() OVER (PARTITION BY reg ORDER BY id) g_f
                              FROM TEMP_temperature_level_bb);'''

    sql_query26 = '''CREATE TEMPORARY TABLE TEMP_temperature_bb_1 (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query26_1 = '''INSERT INTO TEMP_temperature_bb_1 (group_id, reg)
                            SELECT group_id,reg
                              FROM
                           (SELECT
                                  id,
                                  group_id,
                                  reg,
                                  row_number() OVER (PARTITION BY reg ORDER BY id) g_f
                             FROM TEMP_temperature_bb);'''

    sql_query27 = '''CREATE TEMPORARY TABLE TEMP_temperature_duw_1 (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query27_1 = '''INSERT INTO TEMP_temperature_duw_1 (group_id, reg)
                            SELECT group_id,reg
                              FROM
                           (SELECT
                                   id,
                                   group_id,
                                   reg,
                                   row_number() OVER (PARTITION BY reg ORDER BY id) g_f
                              FROM TEMP_temperature_duw);'''

    sql_query28 = '''CREATE TEMPORARY TABLE TEMP_temperature_tcu_1 (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query28_1 = '''INSERT INTO TEMP_temperature_tcu_1 (group_id, reg)
                            SELECT group_id,reg
                              FROM
                           (SELECT
                                   id,
                                   group_id,
                                   reg,
                                   row_number() OVER (PARTITION BY reg ORDER BY id) g_f
                              FROM TEMP_temperature_tcu);'''

    sql_query29 = '''CREATE TEMPORARY TABLE TEMP_temperature_dus_1 (id INTEGER PRIMARY KEY AUTOINCREMENT,group_id INT,reg TEXT);'''

    sql_query29_1 = '''INSERT INTO TEMP_temperature_dus_1 (group_id, reg)
                            SELECT group_id,reg
                              FROM
                           (SELECT
                                   id,
                                   group_id,
                                   reg,
                                   row_number() OVER (PARTITION BY reg ORDER BY id) g_f
                              FROM TEMP_temperature_dus);'''
    
    sql_query19 = '''DROP TABLE IF EXISTS final_table_1;'''

    sql_query20 = '''CREATE TABLE IF NOT EXISTS final_table_1
                        AS
                    SELECT final_table.id,
                           final_table.bs, 
                           final_table.reg,
                           final_table.enm,
                           final_table.ne,
                           final_table.du_type,
                           final_table.temperature,
                           final_table.temperature_level,
                           TEMP_temperature.id as g_f_temperature,
                           TEMP_temperature_level_bb.id as g_f_temperature_level,
                           TEMP_temperature_bb.id as g_f_temperature_bb,
                           TEMP_temperature_duw.id as g_f_temperature_duw,
                           TEMP_temperature_tcu.id as g_f_temperature_tcu,
                           TEMP_temperature_dus.id as g_f_temperature_dus,
                           TEMP_grapf_coloring_result.color_id as color
                      FROM final_table
                 LEFT JOIN TEMP_temperature
                        ON final_table.group_id = TEMP_temperature.group_id 
                 LEFT JOIN TEMP_temperature_level_bb
                        ON final_table.group_id = TEMP_temperature_level_bb.group_id
                 LEFT JOIN TEMP_temperature_bb
                        ON final_table.group_id = TEMP_temperature_bb.group_id
                 LEFT JOIN TEMP_temperature_duw
                        ON final_table.group_id = TEMP_temperature_duw.group_id
                 LEFT JOIN TEMP_temperature_tcu
                        ON final_table.group_id = TEMP_temperature_tcu.group_id
                LEFT JOIN TEMP_temperature_dus
                        ON final_table.group_id = TEMP_temperature_dus.group_id
                 LEFT JOIN TEMP_grapf_coloring_result
                        ON final_table.group_id = TEMP_grapf_coloring_result.group_id
                  ORDER BY TEMP_temperature.id;'''

    sql_query21 = '''SELECT * FROM final_table_1'''

    sql_query22 = '''CREATE TEMPORARY TABLE TEMP_grapf_coloring
                                         AS
                                     SELECT DISTINCT
                                            final_table.group_id,
                                            TEMP_temperature.id as g_f_temperature,
                                            TEMP_temperature_level_bb.id as g_f_temperature_level,
                                            TEMP_temperature_bb.id as g_f_temperature_bb,
                                            TEMP_temperature_duw.id as g_f_temperature_duw,
                                            TEMP_temperature_tcu.id as g_f_temperature_tcu,
                                            TEMP_temperature_dus.id as g_f_temperature_dus,
    										TEMP_temperature_1.id as g_f_temperature_1,
                                            TEMP_temperature_level_bb_1.id as g_f_temperature_level_1,
                                            TEMP_temperature_bb_1.id as g_f_temperature_bb_1,
                                            TEMP_temperature_duw_1.id as g_f_temperature_duw_1,
                                            TEMP_temperature_tcu_1.id as g_f_temperature_tcu_1,
                                            TEMP_temperature_dus_1.id as g_f_temperature_dus_1
                                       FROM final_table
                                  LEFT JOIN TEMP_temperature
                                         ON final_table.group_id = TEMP_temperature.group_id 
                                  LEFT JOIN TEMP_temperature_level_bb
                                         ON final_table.group_id = TEMP_temperature_level_bb.group_id
                                  LEFT JOIN TEMP_temperature_bb
                                         ON final_table.group_id = TEMP_temperature_bb.group_id
                                  LEFT JOIN TEMP_temperature_duw
                                         ON final_table.group_id = TEMP_temperature_duw.group_id
                                  LEFT JOIN TEMP_temperature_tcu
                                         ON final_table.group_id = TEMP_temperature_tcu.group_id
                                  LEFT JOIN TEMP_temperature_dus
                                         ON final_table.group_id = TEMP_temperature_dus.group_id
                                      LEFT JOIN TEMP_temperature_1
                                         ON final_table.group_id = TEMP_temperature_1.group_id 
                                  LEFT JOIN TEMP_temperature_level_bb_1
                                         ON final_table.group_id = TEMP_temperature_level_bb_1.group_id
                                  LEFT JOIN TEMP_temperature_bb_1
                                         ON final_table.group_id = TEMP_temperature_bb_1.group_id
                                  LEFT JOIN TEMP_temperature_duw_1
                                         ON final_table.group_id = TEMP_temperature_duw_1.group_id
                                  LEFT JOIN TEMP_temperature_tcu_1
                                         ON final_table.group_id = TEMP_temperature_tcu_1.group_id
                                    LEFT JOIN TEMP_temperature_dus_1
                                         ON final_table.group_id = TEMP_temperature_dus_1.group_id;'''

    sql_query22_1 = '''SELECT * FROM TEMP_grapf_coloring;'''

    sql_query23 = '''CREATE TEMPORARY TABLE TEMP_grapf_coloring_result (group_id INT, color_id INT);'''

    sql_query23_1 = '''INSERT INTO TEMP_grapf_coloring_result (group_id,color_id)
                            VALUES (:group_id,:color_id);'''

    if len(db_list) > 0:
        try:
            db_connection = sqlite3.connect('1.db')
            db_cursor = db_connection.cursor()
            
            db_cursor.execute(sql_query1)
            db_connection.commit()

            db_cursor.execute(sql_query2.format(datetime_str))
            db_connection.commit()
            answer = db_cursor.fetchone()
            
            if answer is None:

                db_cursor.execute(sql_query3.format(datetime_str))
                db_connection.commit()

                db_cursor.execute(sql_query2.format(datetime_str))
                db_connection.commit()

                date_time_id = db_cursor.fetchone()[0]

                db_cursor.execute(sql_query4)
                db_connection.commit()

                db_cursor.executemany(sql_query5.format(date_time_id),db_list)
                db_connection.commit()

                db_cursor.execute(sql_query6)
                db_connection.commit()

                db_cursor.execute(sql_query6_1.format(date_time_id))
                db_connection.commit()

                db_cursor.execute(sql_query7)
                db_connection.commit()

                db_cursor.execute(sql_query7_1.format(date_time_id))
                db_connection.commit()

                db_cursor.execute(sql_query8_1)
                db_connection.commit()

                db_cursor.execute(sql_query8.format(date_time_id))
                db_connection.commit()

                db_cursor.execute(sql_query9)
                db_connection.commit()

                db_cursor.execute(sql_query10)
                db_connection.commit()

                db_cursor.execute(sql_query11)
                db_connection.commit()

                db_cursor.execute(sql_query12_1)
                db_connection.commit()

                db_cursor.execute(sql_query12_2)
                db_connection.commit()

                db_cursor.execute(sql_query13)
                db_connection.commit()

                db_cursor.execute(sql_query14_1)
                db_connection.commit()

                db_cursor.execute(sql_query14_2)
                db_connection.commit()

                db_cursor.execute(sql_query15)
                db_connection.commit()

                db_cursor.execute(sql_query16_1)
                db_connection.commit()

                db_cursor.execute(sql_query16_2)
                db_connection.commit()

                db_cursor.execute(sql_query17_1)
                db_connection.commit()

                db_cursor.execute(sql_query17_2)
                db_connection.commit()

                db_cursor.execute(sql_query17_3)
                db_connection.commit()
                
                db_cursor.execute(sql_query18_1)
                db_connection.commit()

                db_cursor.execute(sql_query18_2)
                db_connection.commit()

                db_cursor.execute(sql_query18_3)
                db_connection.commit()

                db_cursor.execute(sql_query24)
                db_connection.commit()

                db_cursor.execute(sql_query24_1)
                db_connection.commit()

                db_cursor.execute(sql_query25)
                db_connection.commit()

                db_cursor.execute(sql_query25_1)
                db_connection.commit()

                db_cursor.execute(sql_query26)
                db_connection.commit()

                db_cursor.execute(sql_query26_1)
                db_connection.commit()

                db_cursor.execute(sql_query27)
                db_connection.commit()

                db_cursor.execute(sql_query27_1)
                db_connection.commit()

                db_cursor.execute(sql_query28)
                db_connection.commit()

                db_cursor.execute(sql_query28_1)
                db_connection.commit()

                db_cursor.execute(sql_query29)
                db_connection.commit()

                db_cursor.execute(sql_query29_1)
                db_connection.commit()

                db_cursor.execute(sql_query19)
                db_connection.commit()
                
                db_cursor.execute(sql_query22)
                db_connection.commit()

                db_cursor.execute(sql_query22_1)
                db_connection.commit()

                grapf_coloring_list = list()

                counter = 0
                for result in db_cursor:
                    counter = counter + 1 
                    temp_list = list()
                    for sout in result:
                        if str(sout).isdigit(): 
                            temp_list.append(int(str(sout)))
                    if len(temp_list) == 13:    
                        grapf_coloring_list.append(temp_list)

                if len(grapf_coloring_list) != counter:
                    print("grapf_coloring: не все группы обработаны. len(grapf_coloring):{}, counter:{}".format(len(grapf_coloring_list),counter))
                else:
                    print("grapf_coloring: все группы обработаны. len(grapf_coloring):{}, counter:{}".format(len(grapf_coloring_list),counter))
                
                grapf_coloring_result_dict = list()
                grapf_coloring_result_dict = grapf_coloring(grapf_coloring_list)

                db_cursor.execute(sql_query23)
                db_connection.commit()

                db_cursor.executemany(sql_query23_1,grapf_coloring_result_dict)
                db_connection.commit()

                #--------------------------------

                db_cursor.execute(sql_query20)
                db_connection.commit()

                db_cursor.execute(sql_query21)
                db_connection.commit()
                
                for result in db_cursor:
                    temp_list = list()
                    for sout in result:
                        temp_list.append(str(sout))
                    result_list.append(temp_list)
            
            else:
                print(datetime_str + " alredy exist")

        except sqlite3.Error as error:
            result_error_list.append("db_main_error " + str(e))

        finally:
            if (db_connection):
                db_connection.close()
                print("Подключение к БД закрыто")
    return result_list

def tcu_to_db(tcu_list: list) -> list:

    result_list = list()

    sql_query1 = '''DROP TABLE IF EXISTS tcu_ip;'''

    sql_query2 = '''CREATE TABLE IF NOT EXISTS tcu_ip (
                                                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                  enm TEXT NOT NULL,
                                                   ne TEXT NOT NULL,
                                                   ip TEXT NOT NULL);'''

    sql_query3 = '''INSERT INTO tcu_ip (
                                   enm,
                                    ne,
                                    ip)
                                 VALUES (:enm,:ne,:ip);'''

    sql_query4 = '''SELECT enm,ne,ip FROM tcu_ip'''

    try:
        db_connection = sqlite3.connect('1.db')
        db_cursor = db_connection.cursor()
        
        if tcu_list == []:

            db_cursor.execute(sql_query4)
            db_connection.commit()
            
            for result in db_cursor:
                temp_list = list()
                for sout in result:
                    temp_list.append(str(sout))

                if len(temp_list) == 3:
                    temp_dict = dict() 
                    temp_dict["enm"] = temp_list[0]
                    temp_dict["ne"] = temp_list[1]
                    temp_dict["ip"] = temp_list[2]
                    result_list.append(temp_dict)
        else:

            db_cursor.execute(sql_query1)
            db_connection.commit()

            db_cursor.execute(sql_query2)
            db_connection.commit()

            db_cursor.executemany(sql_query3,tcu_list)
            db_connection.commit()

            db_cursor.execute(sql_query4)
            db_connection.commit()
            
            for result in db_cursor:
                temp_list = list()
                for sout in result:
                    temp_list.append(str(sout))
                    if len(temp_list) == 3:
                        temp_dict = dict() 
                        temp_dict["enm"] = temp_list[0]
                        temp_dict["ne"] = temp_list[1]
                        temp_dict["ip"] = temp_list[2]
                        result_list.append(temp_dict)

    except sqlite3.Error as error:
        result_error_list.append("db_tcu_error " + str(e))

    finally:
        if (db_connection):
            db_connection.close()
            print("Подключение к БД закрыто")
        
    return result_list

def get_tcu_ip_from_enm() -> list:

    result_list = list()
    try:
        enm_sftp_connection_dict = enm_connection_information()
        
        for enm,connection_data in enm_sftp_connection_dict.items():

            enm_session = enmscripting.open(connection_data["url"],connection_data["username"],connection_data["password"])
            cmd = enm_session.command()
            command = "cmedit get * STNConnectivityInformation.ipAddress -s -t"

            response = cmd.execute(command)
            if len(response.get_output().groups()) > 0:
                table = response.get_output().groups()[0]
                
                for line in table:
                    tmp_list = list()
                    tmp_list.append(enm)
                    for cell in line:
                        tmp_list.append(cell.value())
                    
                    result_list.append("|".join(tmp_list))

            enmscripting.close(enm_session)

    except Exception as e:
        result_error_list.append("enm_tcu_error "+ str(e)) 

    return result_list

def tcu_parser(tcu_list:list) -> list:
    result_list = list()
    for tcu_str in tcu_list:

        temp_list = tcu_str.split("|")
        tcu_dict = dict()

        if len(temp_list) == 5 and temp_list[2] == "SYNCHRONIZED":
            
            tcu_dict["enm"] = temp_list[0]
            tcu_dict["ne"] = temp_list[1]
            tcu_dict["ip"] = temp_list[4]

            result_list.append(tcu_dict)

    return result_list

def paramiko_get_data(ip) -> str:
    result = str()
    try:
        session = paramiko.SSHClient()
        session.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        session.connect(hostname=ip.replace(" ","").replace("\n",""),username="admin",password="hidden",look_for_keys=False,allow_agent=False,auth_timeout=True)

        with session.invoke_shell() as ssh:
            time.sleep(2)
            ssh.send("gettemperature\n")
            time.sleep(0.5)
            result = ssh.recv(1000).decode("utf-8")

    except Exception as e:
        #result_error_list.append("paramiko_tcu_error ip: " + str(ip) + " " + str(e))
        print(e)
    finally:
        session.close()

    return result

def get_tcu_temp(inter_dict:dict) -> dict:

    if len(inter_dict) == 3:

        tcu_dict = dict()
        tcu_dict["enm"] = inter_dict["enm"]
        tcu_dict["du_type"] = "TCU"
        tcu_dict["ne"] = inter_dict["ne"]

        result = paramiko_get_data(inter_dict["ip"])

        if result != "":
            result_list = result.split("\n")
            for item in result_list:
                find_1 = re.search(r"Temp sensor 1",item)
                if find_1:
                    find_2 = re.search(r"[-\d]+ C",item)
                    if find_2:
                        tcu_dict["temperature"] = find_2.group(0)[0:-2]

        tcu_dict["temperature_level"] = None

    return tcu_dict

def get_tcu_temp_in_threads(threads_count) -> list:

    tcu_ip_list = tcu_to_db(tcu_parser(get_tcu_ip_from_enm()))

    #print("Получен список из {} TCU".format(len(tcu_ip_list)))

    with ThreadPoolExecutor(max_workers = threads_count) as executor:
        tcu_result_generator = executor.map(get_tcu_temp, tcu_ip_list)

    tcu_result_list = list()

    for item in tcu_result_generator:
        tcu_result_list.append(item)

    return tcu_result_list

def reg_to_db(reg_list: list):

    sql_query1 = '''CREATE TEMPORARY TABLE TEMP_bs_reg (bs TEXT NOT NULL, reg TEXT NOT NULL);'''

    sql_query2 = '''INSERT INTO TEMP_bs_reg (bs,reg)
                            VALUES (:bs,:reg);'''

    sql_query3 = '''CREATE TABLE IF NOT EXISTS bs_reg (
                                                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                   bs TEXT NOT NULL,
                                                  reg TEXT NOT NULL);'''

    sql_query4 = '''INSERT INTO bs_reg (bs, reg)
                                   SELECT TEMP3.bs, TEMP3.reg
                                     FROM
                                          (SELECT TEMP2.bs, TEMP2.reg, bs_reg.id
                                             FROM 
                                                  (SELECT DISTINCT bs, reg
                                                     FROM TEMP_bs_reg) AS TEMP2
                                        LEFT JOIN bs_reg
                                               ON TEMP2.bs = bs_reg.bs AND TEMP2.reg = bs_reg.reg) AS TEMP3
                                    WHERE TEMP3.id IS NULL;'''
    try:
        db_connection = sqlite3.connect('1.db')
        db_cursor = db_connection.cursor()
        
        if reg_list == []:

            db_cursor.execute(sql_query3)
            db_connection.commit()
            
        else:
            db_cursor.execute(sql_query1)
            db_connection.commit()

            db_cursor.executemany(sql_query2,reg_list)
            db_connection.commit()

            db_cursor.execute(sql_query3)
            db_connection.commit()

            db_cursor.execute(sql_query4)
            db_connection.commit()

    except sqlite3.Error as error:
        result_error_list.append("bs_reg_db_error "+ str(e)) 

    finally:
        if (db_connection):
            db_connection.close()
            print("Подключение к БД закрыто")
        
def bs_reg_update():

    result_list = list()

    target_path = "L:\\technical\\Развитие\\БС_М\\БС_М.xlsx"

    destination_dir = os.getcwd() + "\\temp\\"
    destination_file = "temp.xlsx"
    destination_path = destination_dir + destination_file
    try:
        if not os.path.isdir(destination_dir):
            os.makedirs(destination_dir)

        if os.path.exists(target_path):
            shutil.copy(target_path,destination_path)

        if os.path.exists(destination_path):
            wb = openpyxl.load_workbook(destination_path)

            ws = wb.active
            m_row = ws.max_row
            for i in range(1, m_row + 1):
                temp_dict = dict()
                
                cell_obj = ws.cell(row = i, column = 1).value
                find_1 = re.search(r"[MSCOR]{2}\d{4}",cell_obj)
                if find_1:
                    temp_dict["bs"] = find_1.group(0)

                cell_obj = ws.cell(row = i, column = 10).value
                if cell_obj == "Центр":
                    temp_dict["reg"] = "CNT"
                elif cell_obj == "СВ":
                    temp_dict["reg"] = "NEA"
                elif cell_obj == "ЮВ":
                    temp_dict["reg"] = "SEA"
                elif cell_obj == "СЗ":
                    temp_dict["reg"] = "NWS"
                elif cell_obj == "ЮЗ":
                    temp_dict["reg"] = "SWS"

                if len(temp_dict) == 2:
                    result_list.append(temp_dict)

    except Exception as e:
        result_error_list.append("bs_reg_error "+ str(e)) 
    finally:
        pass
        #shutil.rmtree(destination_dir,ignore_errors=True)

    reg_to_db(result_list)

def openpyxl_xlsx_create(result_table,dt:datetime.datetime):

    max_number = 100000

    wb = openpyxl.Workbook()
    ws = wb.active
    
    template_datetime_str = "{y1}{m1}{d1}.{H1}{M1}_{H2}{M2}"

    datetime_str = get_path_date(dt,template_datetime_str)

    ws.title = datetime_str

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 7
    ws.column_dimensions["G"].width = 7
    ws.column_dimensions["H"].width = 7
    ws.column_dimensions["I"].width = 7
    ws.column_dimensions["J"].width = 7
    ws.column_dimensions["K"].width = 7
    ws.column_dimensions["L"].width = 7
    ws.column_dimensions["M"].width = 7
    ws.column_dimensions["N"].width = 7
    ws.row_dimensions[1].height = 115

    alignment = Alignment(horizontal='left',vertical='bottom',text_rotation=90,wrap_text=False,shrink_to_fit=False,indent=0)
    font = Font(name='Calibri',size=11,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
    border = Border(left=Side(border_style="thin", color='FF000000'),right=Side(border_style="thin", color='FF000000'),top=Side(border_style="thin", color='FF000000'),bottom=Side(border_style="thin", color='FF000000'))
    
    colors_dist = {1 : PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type = "solid"),
                   2 : PatternFill(start_color="00FFFF99", end_color="00FFFF99", fill_type = "solid"),
                   3 : PatternFill(start_color="0099FF99", end_color="0099FF99", fill_type = "solid"),
                   4 : PatternFill(start_color="0092CDDC", end_color="0092CDDC", fill_type = "solid"),
                   5 : PatternFill(start_color="00FABF8F", end_color="00FABF8F", fill_type = "solid"),
                   6 : PatternFill(start_color="00CCCAFF", end_color="00CCCAFF", fill_type = "solid"),
                   7 : PatternFill(start_color="00FF99CC", end_color="00FF99CC", fill_type = "solid"),
                   8 : PatternFill(start_color="0066FFFF", end_color="0066FFFF", fill_type = "solid"),
                   9 : PatternFill(start_color="0000FA9A", end_color="0000FA9A", fill_type = "solid"),
                   10 : PatternFill(start_color="00FFF1B7", end_color="00FFF1B7", fill_type = "solid"),
                   11 : PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type = "solid"),
                   12 : PatternFill(start_color="00FF0066", end_color="00FF0066", fill_type = "solid"),
                   13 : PatternFill(start_color="00CC66FF", end_color="00CC66FF", fill_type = "solid"),
                   14 : PatternFill(start_color="00FFCCCC", end_color="00FFCCCC", fill_type = "solid"),
                   15 : PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type = "solid"),
                   16 : PatternFill(start_color="00CCCC00", end_color="00CCCC00", fill_type = "solid"),
                   17 : PatternFill(start_color="0033CCCC", end_color="0033CCCC", fill_type = "solid"),
                   18 : PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type = "solid"),
                   19 : PatternFill(start_color="00FFCC66", end_color="00FFCC66", fill_type = "solid"),
                   20 : PatternFill(start_color="00FF7C80", end_color="00FF7C80", fill_type = "solid"),
                   21 : PatternFill(start_color="00DDDDDD", end_color="00DDDDDD", fill_type = "solid"),
                   22 : PatternFill(start_color="004BACC6", end_color="004BACC6", fill_type = "solid"),
                   23 : PatternFill(start_color="009BBB59", end_color="009BBB59", fill_type = "solid"),
                   24 : PatternFill(start_color="00F79646", end_color="00F79646", fill_type = "solid"),
                   25 : PatternFill(start_color="004F81BD", end_color="004F81BD", fill_type = "solid")}

    ws.cell(1,1).value = "id"
    ws.cell(1,2).value = "bs"
    ws.cell(1,3).value = "reg"
    ws.cell(1,4).value = "enm"
    ws.cell(1,5).value = "ne"
    ws.cell(1,6).value = "du_type"
    ws.cell(1,7).value = "temperature"
    ws.cell(1,8).value = "temperature_level"
    ws.cell(1,9).value = "g_f_temperature"
    ws.cell(1,10).value = "g_f_temperature_level"
    ws.cell(1,11).value = "g_f_temperature_bb"
    ws.cell(1,12).value = "g_f_temperature_duw"
    ws.cell(1,13).value = "g_f_temperature_tsu"
    ws.cell(1,14).value = "g_f_temperature_dus"

    for i in range(14):
        i = i + 1
        ws.cell(1,i).font = font
        ws.cell(1,i).alignment = alignment
        ws.cell(1,i).border = border


    i = 1
    for line in result_table:
        i = i + 1
        j = 0
        for cell in line[:-1]:
            j = j + 1

            ws.cell(i,j).fill = colors_dist[int(line[-1])]
            ws.cell(i,j).border = border
            
            if j in [1,7,8]:
                if cell.isdigit():
                    ws.cell(i,j).value = int(cell)
            elif j in [9,10,11,12,13,14]:
                if cell.isdigit():
                    ws.cell(i,j).value = int(cell) + max_number
            else: 
                ws.cell(i,j).value = cell

    ws.auto_filter.ref = ws.dimensions

    img = openpyxl.drawing.image.Image
    img1 = img("description.JPG")
    ws.add_image(img1,"P2")


    report_path = get_report_path(dt,"")

    template_datetime_str = "bs_temperature_{y1}{m1}{d1}_{H1}{M1}_{H2}{M2}.xlsx"

    datetime_str = get_path_date(dt,template_datetime_str)
    print(report_path + datetime_str)
    wb.save(report_path + datetime_str)
    
    report_path = "L:\\technical\\RNC_ALL\\bs_temperature\\" + dt.strftime("%Y") + "\\" + dt.strftime("%m") + "\\" + dt.strftime("%d") + "\\"

    if not os.path.isdir(report_path):
        os.makedirs(report_path)
        print(report_path + datetime_str)
        wb.save(report_path + datetime_str)
    else:
        print(report_path + datetime_str)
        wb.save(report_path + datetime_str)  

def grapf_coloring(grapf_coloring_list:list) -> list:
    colors_count = 0 

    sort_dict = dict()
    max_sort_item = 0
    for items in grapf_coloring_list:
        for i,item in enumerate(items):
            if i == 0:
                group_id = item
            else:
                if i > colors_count:
                    colors_count = i
                if i not in sort_dict:
                    sort_dict[i] = dict()
                sort_dict[i][item] = group_id
                if item > max_sort_item:
                    max_sort_item = item

    colors_index_list = [i+1 for i in range(colors_count*2+1)]

    neibor_dict = dict()            
    for items in grapf_coloring_list:
        for i,item in enumerate(items):
            if i == 0:
                group_id = item
                neibor_dict[group_id] = dict()
                neibor_dict[group_id]["color"] = 0
                neibor_dict[group_id]["avaliable_color"] = set(colors_index_list)
                neibor_dict[group_id]["neibor_count"] = 0
                neibor_dict[group_id]["neibor_set"] = set()
            else:
                if item == 1:
                    neibor_down = sort_dict[i][item + 1]
                    neibor_dict[group_id]["neibor_set"].add(neibor_down)
                    neibor_dict[group_id]["neibor_count"] = len(neibor_dict[group_id]["neibor_set"])

                elif item == max_sort_item:
                    neibor_up = sort_dict[i][item - 1]
                    neibor_dict[group_id]["neibor_set"].add(neibor_up)
                    neibor_dict[group_id]["neibor_count"] = len(neibor_dict[group_id]["neibor_set"])

                else:
                    neibor_up = sort_dict[i][item - 1]
                    neibor_down = sort_dict[i][item + 1]
                    neibor_dict[group_id]["neibor_set"].add(neibor_up)
                    neibor_dict[group_id]["neibor_set"].add(neibor_down)
                    neibor_dict[group_id]["neibor_count"] = len(neibor_dict[group_id]["neibor_set"])

    neibor_count_sort_dict = dict()
    for key,value in neibor_dict.items():
        neibor_count = value["neibor_count"]

        if neibor_count not in neibor_count_sort_dict:
            neibor_count_sort_dict[neibor_count] = list()
        
        neibor_count_sort_dict[neibor_count].append(key)

    neibor_count_iter_list = sorted(neibor_count_sort_dict.keys(),reverse=True)

    for iterator in neibor_count_iter_list:
        group_id_list = neibor_count_sort_dict[iterator]
        for group_id in group_id_list:

            for neibor_group_id in neibor_dict[group_id]["neibor_set"]:
                neibor_dict[group_id]["avaliable_color"].discard(neibor_dict[neibor_group_id]["color"])
            color_index = neibor_dict[group_id]["avaliable_color"].pop()

            neibor_dict[group_id]["color"] = color_index

    grapf_coloring_result_list = list()
    for key,value in neibor_dict.items():
        temp = dict()
        temp["group_id"] = key
        temp["color_id"] = neibor_dict[key]["color"]
        grapf_coloring_result_list.append(temp)

    return grapf_coloring_result_list

def get_dus_ip_from_enm() -> list:

    result_list = list()
    try:
        enm_sftp_connection_dict = enm_connection_information()
        
        for enm,connection_data in enm_sftp_connection_dict.items():

            enm_session = enmscripting.open(connection_data["url"],connection_data["username"],connection_data["password"])
            cmd = enm_session.command()
            command = "cmedit get * IpHostLink.ipv4Addresses --netype=ERBS -s -t"

            response = cmd.execute(command)
            if len(response.get_output().groups()) > 0:
                table = response.get_output().groups()[0]
                
                for line in table:
                    tmp_list = list()
                    tmp_list.append(enm)
                    for cell in line:
                        tmp_list.append(cell.value())
                    
                    result_list.append("|".join(tmp_list))

            enmscripting.close(enm_session)

    except Exception as e:
        result_error_list.append("enm_dus_error "+ str(e)) 

    return result_list

def dus_parser(dus_list:list) -> list:
    result_list = list()
    for dus_str in dus_list:

        temp_list = dus_str.split("|")
        dus_dict = dict()

        if len(temp_list) == 7 and temp_list[2] == "SYNCHRONIZED":
            
            dus_dict["enm"] = temp_list[0]
            dus_dict["ne"] = temp_list[1]
            dus_dict["ip"] = temp_list[6][1:-1]

            result_list.append(dus_dict)

    return result_list

def dus_to_db(dus_list: list) -> list:

    result_list = list()

    sql_query1 = '''DROP TABLE IF EXISTS dus_ip;'''

    sql_query2 = '''CREATE TABLE IF NOT EXISTS dus_ip (
                                                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                  enm TEXT NOT NULL,
                                                   ne TEXT NOT NULL,
                                                   ip TEXT NOT NULL);'''

    sql_query3 = '''INSERT INTO dus_ip (
                                   enm,
                                    ne,
                                    ip)
                                 VALUES (:enm,:ne,:ip);'''

    sql_query4 = '''SELECT enm,ne,ip FROM dus_ip'''

    try:
        db_connection = sqlite3.connect('1.db')
        db_cursor = db_connection.cursor()
        
        if dus_list == []:

            db_cursor.execute(sql_query4)
            db_connection.commit()
            
            for result in db_cursor:
                temp_list = list()
                for sout in result:
                    temp_list.append(str(sout))

                if len(temp_list) == 3:
                    temp_dict = dict() 
                    temp_dict["enm"] = temp_list[0]
                    temp_dict["ne"] = temp_list[1]
                    temp_dict["ip"] = temp_list[2]
                    result_list.append(temp_dict)
        else:

            db_cursor.execute(sql_query1)
            db_connection.commit()

            db_cursor.execute(sql_query2)
            db_connection.commit()

            db_cursor.executemany(sql_query3,dus_list)
            db_connection.commit()

            db_cursor.execute(sql_query4)
            db_connection.commit()
            
            for result in db_cursor:
                temp_list = list()
                for sout in result:
                    temp_list.append(str(sout))
                    if len(temp_list) == 3:
                        temp_dict = dict() 
                        temp_dict["enm"] = temp_list[0]
                        temp_dict["ne"] = temp_list[1]
                        temp_dict["ip"] = temp_list[2]
                        result_list.append(temp_dict)

    except sqlite3.Error as error:
        result_error_list.append("db_dus_error " + str(e)) 

    finally:
        if (db_connection):
            db_connection.close()
            print("Подключение к БД закрыто")
        
    return result_list

def get_dus_temp(inter_dict:dict) -> dict:

    if len(inter_dict) == 3:

        dus_dict = dict()
        dus_dict["enm"] = inter_dict["enm"]
        dus_dict["du_type"] = "DUS"
        dus_dict["ne"] = inter_dict["ne"]

        result = paramiko_get_dus_data(inter_dict["ip"])
        temperature_list = list()
       
        for item in result.replace("\r","\n").split("\n"):

            find_1 = re.search(r"Temp:",item)
            if find_1:
                point1 = find_1.end()
                temp = item[point1:].strip().split(" ")

                for temperature in temp:
                    if temperature.isdigit():
                        temperature_list.append(int(temperature))

        if temperature_list != []:
            dus_dict["temperature"] = int(sum(temperature_list)/len(temperature_list))
            dus_dict["temperature_level"] = None

        """
        for item in result.replace(" ","").split("\n"):
            
            find_1 = re.search(r"\|Temperature\(C\)\|",item)
            if find_1:
                point1 = find_1.end()
                temperature = item[point1:-2]
                if temperature.isdigit():
                    temperature_list.append(int(temperature))

        if temperature_list != []:
            dus_dict["temperature"] = min(temperature_list)
            dus_dict["temperature_level"] = None
        """
    return dus_dict

def get_dus_temp_in_threads(threads_count) -> list:

    dus_ip_list = dus_to_db(dus_parser(get_dus_ip_from_enm()))

    with open("dus_ip_list.txt","w+") as f_out:
        for item in dus_ip_list:
            temp = list()
            for key,value in item.items():
                temp.append(str(value))
            f_out.write("|".join(temp) + "\n")

    #print("Получен список из {} DUS".format(len(dus_ip_list)))

    with ThreadPoolExecutor(max_workers = threads_count) as executor:
        dus_result_generator = executor.map(get_dus_temp, dus_ip_list)

    dus_result_list = list()

    for item in dus_result_generator:
        dus_result_list.append(item)

    return dus_result_list

def paramiko_get_dus_data(ip) -> str:
    result = str()
    try:
        session = paramiko.SSHClient()
        session.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        session.connect(hostname=ip.replace(" ","").replace("\n",""),username="rbs",password="RBSEricsson12#",look_for_keys=False,allow_agent=False,timeout=10)

        with session.invoke_shell() as ssh:
            time.sleep(1.5)
            #ssh.send("sfp -a\n")
            ssh.send("readPower\n")
            time.sleep(2.5)
            result = ssh.recv(1000000).decode("utf-8")

    except Exception as e:
        #result_error_list.append("paramiko_dus_error ip: " + str(ip) + " " + str(e)) 
        print(e)
    finally:
        session.close()

    return result

def get_duw_ip_from_enm() -> list:
    
    result_list = list()
    try:
        enm_sftp_connection_dict = enm_connection_information()
        
        for enm,connection_data in enm_sftp_connection_dict.items():

            enm_session = enmscripting.open(connection_data["url"],connection_data["username"],connection_data["password"])
            cmd = enm_session.command()
            command = "cmedit get * --scopefilter (Slot.productData.(productName=='DUW 31 01')) IpHostLink.ipv4Addresses -s -t"

            response = cmd.execute(command)
            if len(response.get_output().groups()) > 0:
                table = response.get_output().groups()[0]
                
                for line in table:
                    tmp_list = list()
                    tmp_list.append(enm)
                    for cell in line:
                        tmp_list.append(cell.value())
                    
                    result_list.append("|".join(tmp_list))

            enmscripting.close(enm_session)

    except Exception as e:
        result_error_list.append("enm_duw_error "+ str(e)) 

    return result_list

def duw_parser(duw_list:list) -> list:
    result_list = list()
    for duw_str in duw_list:

        temp_list = duw_str.split("|")
        duw_dict = dict()

        if len(temp_list) == 7 and temp_list[2] == "SYNCHRONIZED":
            
            duw_dict["enm"] = temp_list[0]
            duw_dict["ne"] = temp_list[1]
            duw_dict["ip"] = temp_list[6][1:-1]

            result_list.append(duw_dict)

    return result_list

def duw_to_db(duw_list: list) -> list:

    result_list = list()

    sql_query1 = '''DROP TABLE IF EXISTS duw_ip;'''

    sql_query2 = '''CREATE TABLE IF NOT EXISTS duw_ip (
                                                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                  enm TEXT NOT NULL,
                                                   ne TEXT NOT NULL,
                                                   ip TEXT NOT NULL);'''

    sql_query3 = '''INSERT INTO duw_ip (
                                   enm,
                                    ne,
                                    ip)
                                 VALUES (:enm,:ne,:ip);'''

    sql_query4 = '''SELECT enm,ne,ip FROM duw_ip'''

    try:
        db_connection = sqlite3.connect('1.db')
        db_cursor = db_connection.cursor()
        
        if duw_list == []:

            db_cursor.execute(sql_query4)
            db_connection.commit()
            
            for result in db_cursor:
                temp_list = list()
                for sout in result:
                    temp_list.append(str(sout))

                if len(temp_list) == 3:
                    temp_dict = dict() 
                    temp_dict["enm"] = temp_list[0]
                    temp_dict["ne"] = temp_list[1]
                    temp_dict["ip"] = temp_list[2]
                    result_list.append(temp_dict)
        else:

            db_cursor.execute(sql_query1)
            db_connection.commit()

            db_cursor.execute(sql_query2)
            db_connection.commit()

            db_cursor.executemany(sql_query3,duw_list)
            db_connection.commit()

            db_cursor.execute(sql_query4)
            db_connection.commit()
            
            for result in db_cursor:
                temp_list = list()
                for sout in result:
                    temp_list.append(str(sout))
                    if len(temp_list) == 3:
                        temp_dict = dict() 
                        temp_dict["enm"] = temp_list[0]
                        temp_dict["ne"] = temp_list[1]
                        temp_dict["ip"] = temp_list[2]
                        result_list.append(temp_dict)

    except sqlite3.Error as error:
        result_error_list.append("db_duw_error " + str(e)) 

    finally:
        if (db_connection):
            db_connection.close()
            print("Подключение к БД закрыто")
        
    return result_list

def get_duw_temp(inter_dict:dict) -> dict:
    
    duw_dict = dict()
    
    if len(inter_dict) == 3:
        
        duw_dict["enm"] = inter_dict["enm"]
        duw_dict["du_type"] = "DUW"
        duw_dict["ne"] = inter_dict["ne"]

        result = paramiko_get_duw_data(inter_dict["ip"])
        temperature_list = list()
        for item in result.replace("\r","\n").split("\n"):

            find_1 = re.search(r"Temp:",item)
            if find_1:
                point1 = find_1.end()
                temp = item[point1:].strip().split(" ")

                for temperature in temp:
                    if temperature.isdigit():
                        temperature_list.append(int(temperature))

        if temperature_list != []:
            duw_dict["temperature"] = int(sum(temperature_list)/len(temperature_list))
            duw_dict["temperature_level"] = None


    return duw_dict

def get_duw_temp_in_threads(threads_count) -> list:

    duw_ip_list = duw_to_db(duw_parser(get_duw_ip_from_enm()))

    with open("duw_ip_list.txt","w+") as f_out:
        for item in duw_ip_list:
            temp = list()
            for key,value in item.items():
                temp.append(str(value))
            f_out.write("|".join(temp) + "\n")

    #print("Получен список из {} DUW".format(len(duw_ip_list)))

    with ThreadPoolExecutor(max_workers = threads_count) as executor:
        duw_result_generator = executor.map(get_duw_temp, duw_ip_list)

    duw_result_list = list()
    duw_filter_list = dict()
    for item in duw_result_generator:
        duw_result_list.append(item)
        duw_filter_list[item["enm"]+item["ne"]] = 1

    return [duw_result_list,duw_filter_list]

def paramiko_get_duw_data(ip) -> str:
    result = str()
    try:
        session = paramiko.SSHClient()
        session.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        session.connect(hostname=ip.replace(" ","").replace("\n",""),username="rbs",password="RBSEricsson12#",look_for_keys=False,allow_agent=False,timeout=10)

        with session.invoke_shell() as ssh:
            time.sleep(1.5)
            ssh.send("readPower\n")
            time.sleep(2.5)
            result = ssh.recv(1000000).decode("utf-8")

    except Exception as e:
        #result_error_list.append("paramiko_duw_error ip: " + str(ip) + " " + str(e)) 
        print(e)
    finally:
        session.close()

    return result

if __name__ == '__main__' and True:
    
    #----------------------------------------------------------------------------------------------------------------------
    time_control_start_time = datetime.datetime.now()
    time_control_past_time = time_control_start_time
    print("Программа начата. time_control: {}".format(datetime.datetime.now()))  
    #----------------------------------------------------------------------------------------------------------------------
    dt = datetime.datetime.now()

    #файл ошибок
    result_error_list = list()

    try:
        print("Актуализация данных с файла БС_М")

        bs_reg_update()

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Данные bs_reg обновлены. Затрачено времени time_control: {}".format(time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

        print("Получение данных с TCU ...")

        tcu_list = list()
        tcu_list = get_tcu_temp_in_threads(50)

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Получено {} записей. Затрачено времени time_control: {}".format(len(tcu_list),time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

        print("Получение данных с DUS ...")

        dus_list = list()
        dus_list = get_dus_temp_in_threads(50)

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Получено {} записей. Затрачено времени time_control: {}".format(len(dus_list),time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

        print("Получение данных с DUW 3101 ...")

        duw_list = list()
        duw_list, filter_list = get_duw_temp_in_threads(50)

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Получено {} записей. Затрачено времени time_control: {}".format(len(duw_list),time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

        dt = datetime.datetime.now()

        minute_delta = 5

        minutes_in_part = dt.minute % 15

        if minutes_in_part < minute_delta:
            wait_time = minute_delta - minutes_in_part
            print("Задержка до запроса файлов {} минут".format(wait_time))
            time.sleep(wait_time*60)

        dt = datetime.datetime.now()

        print("Подготовка списка файлов и получение их с sftp server ...")

        with open("path_list.txt","w+") as f_out:
            for item in enm_file_get_in_threads(7,dt):
                f_out.write(item + "\n")

        path_list = list()
        with open("path_list.txt","r") as f_in:
            for line in f_in:
                temp_line = line.replace("\n","")
                path_list.append(temp_line)
        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Получено {} файлов. Затрачено времени time_control: {}".format(len(path_list),time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

        print("Обработка файлов ...")
        result_list = list()
        with multiprocessing.Pool(processes = 4) as pool:
            result_list = pool.map(file_execution,path_list)

        #Удаление записей с DUW 3101 которые были получены ранее опросом через paramiko

        bb_duw_list = list()

        for item in result_list:
            if item["enm"] + item["ne"] not in filter_list:
                bb_duw_list.append(item)
            else:
                filter_list[item["enm"] + item["ne"]] = 2

        for key,value in filter_list.items():
            if value == 1:
                print(key)

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Получено записей {} из файлов. Затрачено времени time_control: {}".format(len(result_list),time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

        final_list = bb_duw_list + tcu_list + dus_list + duw_list
        print("Итоговое количество записей в final_list: {}".format(len(final_list)))

        print("Подготовка записей для БД...")
        db_list = list()
        for item in final_list:
            if ("ne" in item) and ("enm" in item) and ("du_type" in item) and ("temperature" in item) and ("temperature_level" in item):
                find_1 = re.search(r"[MSCOR]{2}\d{4,5}",item["ne"])
                if find_1:
                    if len(find_1.group(0)) == 6:
                        item["bs"] = find_1.group(0)
                    elif len(find_1.group(0)) == 7:
                        item["bs"] = find_1.group(0)[0:2] + find_1.group(0)[-4:]

                    if len(item) == 6:
                        db_list.append(item)

        with open("db_list.txt","w+") as f_out:
            for item in db_list:
                f_out.write(str(item) + "\n")    

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Получено корректных записей {}. Затрачено времени time_control: {}".format(len(db_list),time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

        print("Работа с БД ...")
        result_table = addind_to_db(db_list,dt)

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Работа с БД завершена. Затрачено времени time_control: {}".format(time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

        print("Работа с excel ...")

        if result_table != []:
            try:
                openpyxl_xlsx_create(result_table,dt)
            except Exception as e:
                result_error_list.append("openpyxl_error " + str(e))
                print("Проблема с записью в excel")
            finally:
                pass
        else:
            print("result_table не содержит значений")

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Работа с excel завершена. Затрачено времени time_control: {}".format(time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

    except Exception as e:
        result_error_list.append("main_prog " + str(e))

    finally:

        print("Сохранение вспомогательных файлов ...")
        report_path = get_report_path(dt,"")

        with open(report_path + "result_error_list.txt","w+") as f_out:
            for item in result_error_list:
                f_out.write(item + "\n" + "\n")

        with open(report_path + "path_list.txt","w+") as f_out:
            for item in path_list:
                f_out.write(item + "\n")

        with open(report_path + "tcu_out.txt","w+") as f_out:
            for item in tcu_list:
                temp = list()
                for key,value in item.items():
                    temp.append(str(value))
                f_out.write("|".join(temp) + "\n")

        with open(report_path + "dus_out.txt","w+") as f_out:
            for item in dus_list:
                temp = list()
                for key,value in item.items():
                    temp.append(str(value))
                f_out.write("|".join(temp) + "\n")

        with open(report_path + "duw_out.txt","w+") as f_out:
            for item in duw_list:
                temp = list()
                for key,value in item.items():
                    temp.append(str(value))
                f_out.write("|".join(temp) + "\n")

        with open(report_path + "db_list.txt","w+") as f_out:
            for item in db_list:
                temp = list()
                for key,value in item.items():
                    temp.append(str(value))
                f_out.write("|".join(temp) + "\n")

        with open(report_path + "result_table.txt","w+") as f_out:
            for item in result_table:
                temp = list()
                for value in item:
                    temp.append(str(value))
                f_out.write("|".join(temp) + "\n")

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Файлы сохранены. Затрачено времени time_control: {}".format(time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

        print("Удаление временных файлов ...")
        for enm in ["enm7","enm8","enm14"]:
            report_path = get_report_path(dt,enm)
            shutil.rmtree(report_path,ignore_errors=True)

        #----------------------------------------------------------------------------------------------------------------------
        time_control_delte_time = datetime.datetime.now() - time_control_past_time
        time_control_past_time = datetime.datetime.now()
        print("Временные файлы удалены. Затрачено времени time_control: {}".format(time_control_delte_time))   
        #----------------------------------------------------------------------------------------------------------------------

    #----------------------------------------------------------------------------------------------------------------------
    time_control_delte_time = datetime.datetime.now() - time_control_start_time
    print("Программа завершена. time_control: {}".format(datetime.datetime.now()))  
    print("Время работы программы. Total_time_control: {}".format(time_control_delte_time))  
    #----------------------------------------------------------------------------------------------------------------------

if False:
    db_list = list()
    with open("D:\\Romanov\\Script\\sftp\\report\\2023\\04\\19\\2205\\db_list.txt","r") as f:
        for line in f:
            temp = line.replace("\n","").split("|")
            temp_d = dict()
            temp_d["enm"] = temp[0]
            temp_d["du_type"] = temp[1]
            temp_d["ne"] = temp[2]
            temp_d["temperature"] = temp[3]
            temp_d["temperature_level"] = temp[4]
            temp_d["bs"] = temp[5]

            db_list.append(temp_d)

    dt = datetime.datetime.now()
    result_table = addind_to_db(db_list,dt)

    openpyxl_xlsx_create(result_table,dt)

if False:
    grapf_coloring_list = list()
    with open("grapf_coloring_list.txt","r") as f:
        for line in f:
            temp = list()
            for item in line.replace("\n","").split("|"):
                temp.append(int(item))
            grapf_coloring_list.append(temp)

    grapf_coloring_result_dict = list()
    grapf_coloring_result_dict = grapf_coloring(grapf_coloring_list)

    print(grapf_coloring_result_dict[0:1000])

if False:
    path_list = list()
    with open("D:\\bs_temperature\\report\\2023\\05\\12\\0450\\path_list.txt","r") as f_in:
        for line in f_in:
            temp_line = line.replace("\n","")
            path_list.append(temp_line)

    result_list = list()
    for i,item in enumerate(path_list):
        temp_path = item.split("|")
        result_dict = dict()

        if len(temp_path) == 3:
            enm,du_type,path = temp_path

            result_xml = list()
            dict_xml = dict()

            try:
                result_xml = open_gz(path)
            except Exception as e:
                result_list.append("gz_error " + path + " " + str(e))

            if result_xml != []:
                if du_type == "BB":
                    result_dict["enm"] = enm
                    result_dict["du_type"] = du_type
                    try:
                        dict_xml = get_temp_bb_etree(result_xml)
                    except Exception as e:
                        result_list.append("xml_error " + path + " " + str(e))

                elif du_type == "DUW":
                    result_dict["enm"] = enm
                    result_dict["du_type"] = du_type
                    try:
                        dict_xml = get_temp_duw_etree(result_xml)
                    except Exception as e:
                        result_list.append("xml_error " + path + " " + str(e))

                if dict_xml != {}:
                    result_dict.update(get_temp_duw_etree(open_gz(path)))


    with open("expection_0450.txt","w+") as f_out:
        for item in result_list:
            f_out.write(str(item) + "\n")

def test(item: list):
    temp_path = item.split("|")
    result_dict = dict()

    if len(temp_path) == 3:
        enm,du_type,path = temp_path

        result_xml = list()
        dict_xml = dict()

        try:
            result_xml = open_gz(path)
        except Exception as e:
            print("gz_error " + path + " " + str(e))

        if result_xml != []:
            if du_type == "BB":
                result_dict["enm"] = enm
                result_dict["du_type"] = du_type
                try:
                    dict_xml = get_temp_bb_find(result_xml)
                except Exception as e:
                    print("xml_error " + path + " " + str(e))

            elif du_type == "DUW":
                result_dict["enm"] = enm
                result_dict["du_type"] = du_type
                try:
                    dict_xml = get_temp_duw_etree(result_xml)
                except Exception as e:
                    print("xml_error " + path + " " + str(e))

            if dict_xml != {}:
                result_dict.update(dict_xml)

    return result_dict
    
if __name__ == '__main__' and False:
    #----------------------------------------------------------------------------------------------------------------------
    time_control_start_time = datetime.datetime.now()
    time_control_past_time = time_control_start_time
    print("Программа начата. time_control: {}".format(datetime.datetime.now()))  
    #----------------------------------------------------------------------------------------------------------------------
    
    def paramiko_get_data_123(data:list) -> str:
        enm,ne,ip = data
        base = list()
        base.append(ne + " " + ip)
        result = str()
        try:
            session = paramiko.SSHClient()
            session.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            session.connect(hostname=ip.replace(" ","").replace("\n",""),username="admin",password="hidden",look_for_keys=False,allow_agent=False,auth_timeout=True)

            with session.invoke_shell() as ssh:
                time.sleep(2)
                ssh.send("rev\n")
                time.sleep(35)
                """
                ssh.send("calexpdate 20241001\n")
                time.sleep(0.5)
                ssh.send("caldate 20230531\n")
                time.sleep(0.5)
                ssh.send("calvalue -98\n")
                time.sleep(0.5)
                ssh.send("timeservertest start\n")
                time.sleep(0.5)
                ssh.send("getalarmlist\n")
                time.sleep(0.5)
                """
                result = ssh.recv(10000).decode("utf-8")
                temp = result.replace("\r","").split("\n")
                base = base + temp 
                base.append("\n")
        except Exception as e:
            #result_error_list.append("paramiko_tcu_error ip: " + str(ip) + " " + str(e))
            print(e)
        finally:
            session.close()

        return base
    
    l = [["enm","MS0465","10.85.22.225"],
        ["enm","MO5511","10.85.16.253"],
        ["enm","MS4380","10.68.218.169"],
        ["enm","MO1092","10.89.196.245"],
        ["enm","MS0806","10.85.26.181"],
        ["enm","MS0764","10.68.217.173"],
        ["enm","MS3371","10.85.46.229"],
        ["enm","MO2085","10.85.7.229"],
        ["enm","MS3433","10.85.47.121"],
        ["enm","MO0938","10.89.195.221"],
        ["enm","MS2077","10.85.35.141"],
        ["enm","MO5511","10.85.16.253"],
        ["enm","MS0866","10.85.27.121"],
        ["enm","MO3606","10.85.3.137"],
        ["enm","MS4138","10.68.218.141"],
        ["enm","MS4795","10.85.56.181"],
        ["enm","MS1579","10.85.32.69"],
        ["enm","MO6492","10.89.214.37"]]

    base = list()

    with ThreadPoolExecutor(max_workers = 15) as executor:
        result_generator = executor.map(paramiko_get_data_123, l)


    for item in result_generator:
        print(item)


    #----------------------------------------------------------------------------------------------------------------------
    time_control_delte_time = datetime.datetime.now() - time_control_start_time
    print("Программа завершена. time_control: {}".format(datetime.datetime.now()))  
    print("Время работы программы. Total_time_control: {}".format(time_control_delte_time))  
    #----------------------------------------------------------------------------------------------------------------------

if __name__ == '__main__' and False:
    #----------------------------------------------------------------------------------------------------------------------
    time_control_start_time = datetime.datetime.now()
    time_control_past_time = time_control_start_time
    print("Программа начата. time_control: {}".format(datetime.datetime.now()))  
    #----------------------------------------------------------------------------------------------------------------------

    dus_list = list()
    dus_list = get_dus_temp_in_threads(50)
    print(len(dus_list))

    #----------------------------------------------------------------------------------------------------------------------
    time_control_delte_time = datetime.datetime.now() - time_control_start_time
    print("Программа завершена. time_control: {}".format(datetime.datetime.now()))  
    print("Время работы программы. Total_time_control: {}".format(time_control_delte_time))  
    #----------------------------------------------------------------------------------------------------------------------
